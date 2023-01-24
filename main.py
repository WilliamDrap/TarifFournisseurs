from openpyxl import load_workbook, Workbook

liste_onglets = ('AK98', 'WRO 300 H', 'ARTIS & EVOSYS', 'WFR300')

wb_export = Workbook()
wb_export_sheet = wb_export.create_sheet(title='Catalogue')

if __name__ == '__main__':
    workbook = load_workbook(filename="TarifBaxter2023.xlsx")
    w = workbook
    global_row = 2
    for index_onglet in liste_onglets:
        s = w[index_onglet]
        print(index_onglet)
        start_row = 3
        while s.cell(column=1, row=start_row).value is not None:
            wb_export_sheet.cell(column=1, row=global_row, value=index_onglet)
            wb_export_sheet.cell(column=2, row=global_row, value=s.cell(column=1, row=start_row).value)
            wb_export_sheet.cell(column=3, row=global_row, value=s.cell(column=2, row=start_row).value)
            wb_export_sheet.cell(column=4, row=global_row, value=s.cell(column=5, row=start_row).value)
            wb_export_sheet.cell(column=5, row=global_row, value=s.cell(column=6, row=start_row).value)
            start_row += 1
            global_row += 1

    wb_export.save(filename='catalogue_baxter.xlsx')
